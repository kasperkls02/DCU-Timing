import openpyxl
wb = openpyxl.load_workbook('participants.xlsx')
ws = wb.active

ChipCodes = ['TG-65554','RK-83827']
UciID = []

for i in range(1, ws.max_row + 1):
    if str(ws[i][4].value) in ChipCodes:
            print(ws[i][3].value)
            UciID.append(ws[i][3].value)


print(UciID)
